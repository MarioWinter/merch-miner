"""CSV / XLSX parsing helpers shared by admin upload paths and async parser jobs.

Lifted from `admin.py` so worker-side jobs (`parse_bulk_upload_job`) can import
without pulling Django admin into the worker process.

See features/PROJ-25-bulk-asin-scrape-batches.md AC-8 / AC-9 / EC-11 / EC-12.
"""

import csv
import io
import re


ASIN_PATTERN = re.compile(r'^[A-Z0-9]{10}$')


def _parse_uploaded_file(uploaded_file):
    """Parse uploaded CSV or XLSX file into (rows_list, headers_set).

    Detects format via filename extension. xlsx → openpyxl first sheet.
    Returns (rows, headers) or raises ValueError with a user-facing message.

    Numeric ASIN cells (rare, ISBN-style) are zero-padded to 10 chars when the
    column header is exactly 'asin'.
    """
    name = (uploaded_file.name or '').lower()
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook

        try:
            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Could not read Excel file: {exc}") from exc
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel sheet is empty.") from None
        headers_list = [str(h).strip() if h is not None else '' for h in header_row]
        rows = []
        for raw in rows_iter:
            if all(v is None or str(v).strip() == '' for v in raw):
                continue
            row = {}
            for i, value in enumerate(raw):
                if i >= len(headers_list) or not headers_list[i]:
                    continue
                if value is None:
                    cell = ''
                elif isinstance(value, int) and headers_list[i] == 'asin':
                    cell = str(value).zfill(10)
                else:
                    cell = str(value).strip()
                row[headers_list[i]] = cell
            rows.append(row)
        return rows, set(h for h in headers_list if h)

    try:
        decoded = uploaded_file.read().decode('utf-8')
    except UnicodeDecodeError as exc:
        raise ValueError("File is not valid UTF-8.") from exc
    reader = csv.DictReader(io.StringIO(decoded))
    rows = list(reader)
    return rows, set(reader.fieldnames or [])


def normalize_asin_row(row, default_marketplace, asin_pattern=None, valid_marketplaces=None):
    """Normalize one parsed row to a clean dict ready for ScheduledScrapeTarget(...).

    Returns:
        (clean_dict, None) on success — clean_dict has keys 'asin', 'marketplace'.
        (None, error_msg)  on validation failure.

    Args:
        row: a dict produced by `_parse_uploaded_file` (str → str).
        default_marketplace: applied when row has no `marketplace` column.
        asin_pattern: compiled regex; defaults to module ASIN_PATTERN.
        valid_marketplaces: iterable of allowed marketplace values; if None, any
            non-empty marketplace is accepted (parser-level validation skipped).
    """
    if asin_pattern is None:
        asin_pattern = ASIN_PATTERN

    asin = (row.get('asin') or '').strip().upper()
    if not asin:
        return None, "missing asin"
    if not asin_pattern.match(asin):
        return None, f"invalid asin '{asin}'"

    marketplace = (row.get('marketplace') or '').strip() or default_marketplace
    if valid_marketplaces is not None and marketplace not in valid_marketplaces:
        return None, f"invalid marketplace '{marketplace}'"

    return {'asin': asin, 'marketplace': marketplace}, None


def dedupe_within_file(rows_iter):
    """Generator yielding only the latest occurrence per (asin, marketplace).

    Buffers rows into a dict so the *latest* row wins (AC-11). After exhausting
    the input iterator, sets `dedupe_within_file.duplicate_count` on the function
    itself for the caller to read after iteration completes.

    Note: the function attribute is reset at the start of each call so concurrent
    callers in the same process must not share the function reference; in practice
    the parser runs sequentially per batch.
    """
    dedupe_within_file.duplicate_count = 0
    seen = {}
    for row in rows_iter:
        key = (row['asin'], row['marketplace'])
        if key in seen:
            dedupe_within_file.duplicate_count += 1
        seen[key] = row
    yield from seen.values()
