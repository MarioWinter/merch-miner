"""FlyingUpload export service (PROJ-11 Phase S).

PINNED VERSION: FlyingUpload "Excel Standard v2.3" (2023-11-07).
Template stubs: ``publish_app/catalogs/flyingupload_{mba,basic}_template.xlsx``.
See AC-120 for upgrade procedure.

Covers AC-90, AC-92..AC-97, AC-101..AC-107, AC-120, AC-122, AC-127, AC-136.

Exports:
- ``build_mba_bundle(workspace_id, design_ids) -> (zip_bytes, preflight_summary)``
- ``build_basic_bundle(workspace_id, design_ids) -> (zip_bytes, preflight_summary)``
- ``build_mba_csv(workspace_id, design_ids) -> (csv_bytes, preflight_summary)``
- ``build_basic_csv(workspace_id, design_ids) -> (csv_bytes, preflight_summary)``
- ``preflight(workspace_id, design_ids, template, format) -> summary``
"""

from __future__ import annotations

import csv
import io
import logging
import os
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from django.core.files.storage import default_storage

from publish_app.catalogs.flyingupload_maps import (
    FIT_TYPE_MAP,
    FLYINGUPLOAD_PRODUCT_MAP,
    MARKETPLACE_MAP,
    PRICE_COLUMN_ORDER,
    derive_color_mode,
)
from publish_app.models import (
    DesignAsset,
    DesignProductConfig,
    Listing,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Guardrail constants (AC-107).
# ---------------------------------------------------------------------------

MAX_DESIGNS_PER_EXPORT = 500
MAX_ARCHIVE_SIZE_BYTES = 500 * 1024 * 1024  # 500 MB
_XLSX_SIZE_BUDGET = 512 * 1024  # 512 KB reserved for the XLSX itself

TEMPLATES_DIR = (
    Path(__file__).resolve().parent.parent / 'catalogs'
)
MBA_TEMPLATE_PATH = TEMPLATES_DIR / 'flyingupload_mba_template.xlsx'
BASIC_TEMPLATE_PATH = TEMPLATES_DIR / 'flyingupload_basic_template.xlsx'

# Spooled ZIP threshold: above 100 MB swap from in-memory BytesIO to disk.
_SPOOL_MAX_MEM_BYTES = 100 * 1024 * 1024


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class ExportError(Exception):
    """Raised for preflight/guardrail violations. Caller maps to 400."""

    def __init__(
        self,
        code: str,
        message: str = '',
        details: dict[str, Any] | None = None,
    ):
        super().__init__(message or code)
        self.code = code
        self.details = details or {}


# ---------------------------------------------------------------------------
# Internal data classes
# ---------------------------------------------------------------------------


@dataclass
class _SkippedEntry:
    design_id: str
    file_name: str
    reason: str


@dataclass
class _WarningEntry:
    design_id: str
    message: str


@dataclass
class _Summary:
    template: str
    total_designs: int = 0
    ready_rows: int = 0
    skipped: list[dict[str, str]] = field(default_factory=list)
    warnings: list[dict[str, str]] = field(default_factory=list)

    def skip(self, entry: _SkippedEntry) -> None:
        self.skipped.append(
            {
                'design_id': entry.design_id,
                'file_name': entry.file_name,
                'reason': entry.reason,
            },
        )

    def warn(self, entry: _WarningEntry) -> None:
        self.warnings.append(
            {'design_id': entry.design_id, 'message': entry.message},
        )

    def as_dict(self) -> dict[str, Any]:
        return {
            'template': self.template,
            'total_designs': self.total_designs,
            'ready_rows': self.ready_rows,
            'skipped': list(self.skipped),
            'warnings': list(self.warnings),
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _safe_file_name(original: str, asset_uuid: str) -> str:
    """AC-106: collision-safe filename -- ``<stem>-<uuid8>.<ext>``.

    ``asset_uuid`` must be the ``DesignAsset.id`` (str UUID). First 8 hex
    chars are used as a short suffix.
    """
    stem, ext = os.path.splitext(original or 'design.png')
    if not ext:
        ext = '.png'
    suffix = str(asset_uuid).replace('-', '')[:8]
    stem = stem or 'design'
    return f'{stem}-{suffix}{ext}'


def _open_asset_binary(asset: DesignAsset) -> bytes | None:
    """AC-104/AC-105: read a DesignAsset's image bytes.

    Returns None when the binary cannot be fetched (missing file, no
    storage handle). Caller adds the design to ``skipped`` with
    ``image_unavailable`` reason.
    """
    try:
        if getattr(asset, 'file', None) and asset.file:
            return asset.file.read()
    except (OSError, ValueError) as exc:
        logger.warning(
            'flyingupload_export: asset %s file read failed: %s',
            asset.id, exc,
        )
        return None

    file_url = getattr(asset, 'file_url', '') or ''
    if not file_url:
        return None
    if file_url.startswith(('http://', 'https://')):
        # AC-105: cloud assets must have been copied to local storage at
        # import-time. We do NOT call Drive/OneDrive from the export path.
        return None
    try:
        with default_storage.open(file_url, 'rb') as fh:
            return fh.read()
    except (FileNotFoundError, OSError) as exc:
        logger.warning(
            'flyingupload_export: storage open %r failed: %s',
            file_url, exc,
        )
        return None


def _resolve_background_hex(workspace_id: str, design_id: str) -> str:
    """AC-127: BN column = Displate listing's background_color_hex (if any)."""
    hex_val = (
        Listing.objects
        .filter(
            workspace_id=workspace_id,
            design_id=design_id,
            marketplace_type=Listing.MarketplaceType.DISPLATE,
        )
        .values_list('background_color_hex', flat=True)
        .first()
    )
    return hex_val or ''


def _fetch_listings(
    workspace_id: str, design_ids: list[str],
) -> dict[tuple[str, str], Listing]:
    """Preload all Global/MBA/Displate Listings for the targeted designs.

    Returns a dict keyed by ``(design_id, marketplace_type)``.
    """
    qs = (
        Listing.objects
        .filter(
            workspace_id=workspace_id,
            design_id__in=design_ids,
            is_template=False,
        )
        .select_related('idea')
    )
    return {(str(row.design_id), row.marketplace_type): row for row in qs}


def _fetch_designs(
    workspace_id: str, design_ids: list[str],
) -> dict[str, DesignAsset]:
    qs = (
        DesignAsset.objects
        .filter(workspace_id=workspace_id, id__in=design_ids)
        .select_related('collection')
    )
    return {str(d.id): d for d in qs}


def _fetch_product_configs(
    design_ids: list[str],
) -> dict[str, DesignProductConfig]:
    qs = DesignProductConfig.objects.filter(
        design_id__in=design_ids,
        marketplace_type=DesignProductConfig.MarketplaceType.MBA,
    )
    return {str(c.design_id): c for c in qs}


def _get_translation_field(
    listing: Listing | None, lang: str, field: str,
) -> str:
    """Pull ``translations[lang][field]`` with EN / top-level fallback."""
    if listing is None:
        return ''
    translations = listing.translations or {}
    node = translations.get(lang) or {}
    value = node.get(field) or ''
    if value:
        return value
    if lang == 'en':
        return getattr(listing, field, '') or ''
    return ''


def _keywords_csv(listing: Listing | None, lang: str) -> str:
    if listing is None:
        return ''
    keywords = (listing.keywords or {}).get(lang) or []
    return ', '.join(k for k in keywords if isinstance(k, str))


def _disambiguate_file_names(
    designs: dict[str, DesignAsset],
) -> dict[str, str]:
    """AC-106: when two designs share ``file_name``, BOTH get the suffix.

    Returns ``{design_id: safe_file_name}``.
    """
    seen: dict[str, list[str]] = {}
    for did, asset in designs.items():
        seen.setdefault(asset.file_name or 'design.png', []).append(did)

    out: dict[str, str] = {}
    for fname, ids in seen.items():
        if len(ids) == 1:
            # AC-106 still applies when collisions happen -- but a unique
            # filename keeps the raw name (matches the FlyingUpload screenshot
            # in the spec).
            out[ids[0]] = fname
        else:
            for did in ids:
                out[did] = _safe_file_name(fname, did)
    return out


# ---------------------------------------------------------------------------
# MBA row builder
# ---------------------------------------------------------------------------


_MBA_HEADER_COUNT = 66


def _empty_mba_row() -> list[Any]:
    return [''] * _MBA_HEADER_COUNT


def _build_mba_row(
    *,
    safe_file_name: str,
    mba_listing: Listing | None,
    global_listing: Listing | None,
    brand_name: str,
    category: str,
    product_entry: dict,
    workspace_id: str,
    design_id: str,
    summary: _Summary,
) -> list[Any] | None:
    """Build one data row for the MBA XLSX/CSV.

    Returns None when the product key is unknown (row skipped + warning).
    """
    row = _empty_mba_row()

    # A Image Path
    row[0] = f'designs/{safe_file_name}'
    # B intentionally empty
    # C-Q: Title/Description/Tags per lang (DE, FR, IT, ES, JP, EN order for
    # non-EN langs; EN lives at R-T)
    non_en_langs = ('de', 'fr', 'it', 'es', 'ja')
    col = 2  # C (0-indexed)
    for lang in non_en_langs:
        row[col] = _get_translation_field(mba_listing, lang, 'title')
        row[col + 1] = _get_translation_field(mba_listing, lang, 'description')
        row[col + 2] = _keywords_csv(global_listing, lang)
        col += 3
    # R, S, T -- EN
    row[17] = _get_translation_field(mba_listing, 'en', 'title') or (
        mba_listing.title if mba_listing else ''
    )
    row[18] = _get_translation_field(mba_listing, 'en', 'description') or (
        mba_listing.description if mba_listing else ''
    )
    row[19] = _keywords_csv(global_listing, 'en')

    # U Type -- CSV of fit_types for THIS product entry.
    fit_types = product_entry.get('fit_types') or []
    row[20] = ', '.join(fit_types)

    # V Color -- single-value from colors[] via derive_color_mode.
    colors = product_entry.get('colors') or []
    row[21] = derive_color_mode(list(colors))

    # W intentionally empty

    # X-AC Brand duplicated into 6 columns.
    for i in range(23, 29):
        row[i] = brand_name

    # AD-AI Bullet 1 per lang (DE, FR, IT, ES, JP, EN).
    bullet1_langs = ('de', 'fr', 'it', 'es', 'ja', 'en')
    for i, lang in enumerate(bullet1_langs):
        value = _get_translation_field(mba_listing, lang, 'bullet_1')
        if not value and lang == 'en' and mba_listing:
            value = mba_listing.bullet_1 or ''
        row[29 + i] = value

    # AJ-AO Bullet 2 per lang.
    for i, lang in enumerate(bullet1_langs):
        value = _get_translation_field(mba_listing, lang, 'bullet_2')
        if not value and lang == 'en' and mba_listing:
            value = mba_listing.bullet_2 or ''
        row[35 + i] = value

    # AP-AY Color1..Color10.
    for i in range(10):
        if i < len(colors):
            row[41 + i] = colors[i]
    if len(colors) > 10:
        summary.warn(
            _WarningEntry(
                design_id=design_id,
                message=(
                    f'Over 10 colors -- only first 10 exported ('
                    f'{len(colors)} configured)'
                ),
            ),
        )

    # AZ Product.
    product_key = product_entry.get('product_type') or ''
    label = FLYINGUPLOAD_PRODUCT_MAP.get(product_key)
    if not label:
        summary.warn(
            _WarningEntry(
                design_id=design_id,
                message=(
                    f'Unknown product type {product_key!r} -- '
                    f'skipped (EC-48)'
                ),
            ),
        )
        return None
    row[51] = label

    # BA Marketplace CSV.
    marketplaces_entries = product_entry.get('marketplaces') or []
    enabled_codes: list[str] = []
    price_by_mp: dict[str, Any] = {}
    for mp in marketplaces_entries:
        if not isinstance(mp, dict):
            continue
        if not mp.get('enabled'):
            continue
        code = MARKETPLACE_MAP.get(mp.get('marketplace') or '')
        if code:
            enabled_codes.append(code)
            price_by_mp[code] = mp.get('price')
    row[52] = ', '.join(enabled_codes)

    # BB-BH Price per marketplace.
    for idx, (code, _mp) in enumerate(PRICE_COLUMN_ORDER):
        value = price_by_mp.get(code)
        if value in (None, ''):
            row[53 + idx] = ''
        else:
            row[53 + idx] = value

    # BI Print side.
    print_side = product_entry.get('print_side') or 'front'
    if print_side == 'both':
        summary.warn(
            _WarningEntry(
                design_id=design_id,
                message=(
                    '`both` print_side not supported by FlyingUpload v2.3 -- '
                    'exported as `front` (EC-50)'
                ),
            ),
        )
        print_side = 'front'
    row[60] = print_side

    # BJ Draft.
    if mba_listing and mba_listing.publish_mode == Listing.PublishMode.DRAFT:
        row[61] = 'yes'
    else:
        row[61] = ''

    # BK intentionally empty.

    # BL Collection.
    # pull via product_entry-supplied extras
    row[63] = product_entry.get('__collection_name', '') or ''

    # BM Category.
    row[64] = category or ''

    # BN Background color hex (from Displate listing).
    row[65] = _resolve_background_hex(workspace_id, design_id)

    return row


# ---------------------------------------------------------------------------
# Basic row builder
# ---------------------------------------------------------------------------


_BASIC_HEADER_COUNT = 9


def _build_basic_row(
    *,
    safe_file_name: str,
    global_listing: Listing,
) -> list[Any]:
    """Build one data row for the Basic XLSX/CSV. AC-96."""
    row = [''] * _BASIC_HEADER_COUNT
    row[0] = f'designs/{safe_file_name}'
    row[1] = _get_translation_field(global_listing, 'de', 'title')
    row[2] = _get_translation_field(global_listing, 'de', 'description')
    row[3] = _keywords_csv(global_listing, 'de')
    row[4] = (
        _get_translation_field(global_listing, 'en', 'title')
        or (global_listing.title or '')
    )
    row[5] = (
        _get_translation_field(global_listing, 'en', 'description')
        or (global_listing.description or '')
    )
    row[6] = _keywords_csv(global_listing, 'en')

    # Type column: map men->man, women->woman (AC-96).
    type_flags = list(global_listing.type_flags or [])
    row[7] = ', '.join(FIT_TYPE_MAP.get(t, t) for t in type_flags)

    row[8] = global_listing.color_mode or ''
    return row


# ---------------------------------------------------------------------------
# Core MBA pipeline
# ---------------------------------------------------------------------------


def _validate_ids(design_ids: list[str]) -> list[str]:
    """Normalize + enforce AC-107 500-design hard cap."""
    ids = [str(x) for x in (design_ids or []) if x]
    if not ids:
        raise ExportError('design_ids_required', 'At least one design_id is required.')
    if len(ids) > MAX_DESIGNS_PER_EXPORT:
        raise ExportError(
            'max_500_designs_per_export',
            message=(
                f'Max {MAX_DESIGNS_PER_EXPORT} designs per export -- '
                f'got {len(ids)}.'
            ),
        )
    return ids


def _estimate_archive_size(
    designs: dict[str, DesignAsset], ids_in_use: list[str],
) -> tuple[int, list[dict[str, Any]]]:
    total = _XLSX_SIZE_BUDGET
    breakdown: list[dict[str, Any]] = []
    for did in ids_in_use:
        asset = designs.get(did)
        if asset is None:
            continue
        total += int(asset.file_size or 0)
        if (asset.file_size or 0) > 10 * 1024 * 1024:
            breakdown.append(
                {
                    'design_id': did,
                    'file_name': asset.file_name,
                    'file_size': asset.file_size,
                },
            )
    breakdown.sort(key=lambda x: -int(x.get('file_size') or 0))
    return total, breakdown[:10]


def _enforce_size_cap(
    designs: dict[str, DesignAsset], ids_in_use: list[str],
) -> None:
    estimate, breakdown = _estimate_archive_size(designs, ids_in_use)
    if estimate > MAX_ARCHIVE_SIZE_BYTES:
        raise ExportError(
            'estimated_archive_too_large',
            message=(
                f'Estimated archive size {estimate} bytes exceeds '
                f'{MAX_ARCHIVE_SIZE_BYTES} cap.'
            ),
            details={'estimate': estimate, 'top_designs': breakdown},
        )


def _open_template(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(str(path))


def _write_mba_rows_to_sheet(
    ws, rows: list[list[Any]],
) -> None:
    # Data rows start at row 2; template stub already has header row.
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def _pack_zip(
    xlsx_bytes: bytes,
    image_manifest: list[tuple[str, bytes]],
) -> bytes:
    """Pack XLSX + designs/* into a ZIP. XLSX = DEFLATE, images = STORED."""
    # SpooledTemporaryFile keeps small archives in memory, spills to disk.
    spool = tempfile.SpooledTemporaryFile(max_size=_SPOOL_MAX_MEM_BYTES)
    try:
        with zipfile.ZipFile(spool, mode='w', allowZip64=True) as zf:
            # XLSX (compressed).
            zi = zipfile.ZipInfo('flyingupload.xlsx')
            zi.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(zi, xlsx_bytes)
            # designs/* -- stored (no recompression of PNG/JPEG).
            for safe_name, blob in image_manifest:
                zi = zipfile.ZipInfo(f'designs/{safe_name}')
                zi.compress_type = zipfile.ZIP_STORED
                zf.writestr(zi, blob)
        spool.seek(0)
        return spool.read()
    finally:
        spool.close()


def _iterate_mba_context(
    workspace_id: str, design_ids: list[str], summary: _Summary,
):
    """Yield per-row context dicts. Populates ``summary.skipped``.

    Each yielded dict carries the data a downstream writer (XLSX cell / CSV
    cell) needs.
    """
    designs = _fetch_designs(workspace_id, design_ids)
    listings = _fetch_listings(workspace_id, design_ids)
    product_configs = _fetch_product_configs(design_ids)
    safe_names = _disambiguate_file_names(designs)

    # Enforce size cap BEFORE fetching any binaries.
    _enforce_size_cap(designs, list(designs.keys()))

    summary.total_designs = len(design_ids)

    # Preserve caller-provided design_id ordering.
    context_list = []
    image_manifest: list[tuple[str, bytes]] = []
    manifest_seen: set[str] = set()

    for did in design_ids:
        asset = designs.get(did)
        if asset is None:
            summary.skip(
                _SkippedEntry(
                    design_id=did,
                    file_name='',
                    reason='design_not_found',
                ),
            )
            continue

        mba_listing = listings.get(
            (did, Listing.MarketplaceType.MBA),
        )
        global_listing = listings.get(
            (did, Listing.MarketplaceType.GLOBAL),
        )
        if mba_listing is None:
            summary.skip(
                _SkippedEntry(
                    design_id=did,
                    file_name=asset.file_name,
                    reason='no_listing',
                ),
            )
            continue

        config = product_configs.get(did)
        enabled_products: list[dict] = []
        if config is not None:
            for entry in config.products_config or []:
                if isinstance(entry, dict) and entry.get('enabled'):
                    enabled_products.append(entry)
        if not enabled_products:
            summary.skip(
                _SkippedEntry(
                    design_id=did,
                    file_name=asset.file_name,
                    reason='no_enabled_products',
                ),
            )
            continue

        safe_name = safe_names[did]

        # AC-104: fetch binary; skip on failure.
        binary = _open_asset_binary(asset)
        if binary is None:
            summary.skip(
                _SkippedEntry(
                    design_id=did,
                    file_name=asset.file_name,
                    reason='image_unavailable',
                ),
            )
            continue

        if safe_name not in manifest_seen:
            image_manifest.append((safe_name, binary))
            manifest_seen.add(safe_name)

        brand_name = mba_listing.brand_name or ''
        category = mba_listing.category or ''
        collection_name = (
            asset.collection.name if getattr(asset, 'collection', None) else ''
        )

        for entry in enabled_products:
            entry_copy = dict(entry)
            entry_copy['__collection_name'] = collection_name
            context_list.append(
                {
                    'design_id': did,
                    'safe_file_name': safe_name,
                    'mba_listing': mba_listing,
                    'global_listing': global_listing,
                    'brand_name': brand_name,
                    'category': category,
                    'product_entry': entry_copy,
                    'workspace_id': workspace_id,
                },
            )

    return context_list, image_manifest


def _iterate_basic_context(
    workspace_id: str, design_ids: list[str], summary: _Summary,
):
    designs = _fetch_designs(workspace_id, design_ids)
    listings = _fetch_listings(workspace_id, design_ids)
    safe_names = _disambiguate_file_names(designs)
    _enforce_size_cap(designs, list(designs.keys()))

    summary.total_designs = len(design_ids)

    context_list = []
    image_manifest: list[tuple[str, bytes]] = []
    manifest_seen: set[str] = set()

    for did in design_ids:
        asset = designs.get(did)
        if asset is None:
            summary.skip(
                _SkippedEntry(
                    design_id=did,
                    file_name='',
                    reason='design_not_found',
                ),
            )
            continue
        global_listing = listings.get(
            (did, Listing.MarketplaceType.GLOBAL),
        )
        if global_listing is None:
            summary.skip(
                _SkippedEntry(
                    design_id=did,
                    file_name=asset.file_name,
                    reason='no_global_listing',
                ),
            )
            continue

        safe_name = safe_names[did]
        binary = _open_asset_binary(asset)
        if binary is None:
            summary.skip(
                _SkippedEntry(
                    design_id=did,
                    file_name=asset.file_name,
                    reason='image_unavailable',
                ),
            )
            continue
        if safe_name not in manifest_seen:
            image_manifest.append((safe_name, binary))
            manifest_seen.add(safe_name)

        context_list.append(
            {
                'design_id': did,
                'safe_file_name': safe_name,
                'global_listing': global_listing,
            },
        )
    return context_list, image_manifest


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_mba_bundle(
    workspace_id: str, design_ids: list[str],
) -> tuple[bytes, dict[str, Any]]:
    """Build MBA ZIP bundle (AC-101 + AC-102 + AC-103).

    Returns ``(zip_bytes, preflight_summary)``.
    """
    ids = _validate_ids(design_ids)
    summary = _Summary(template='mba')
    contexts, manifest = _iterate_mba_context(workspace_id, ids, summary)

    # Build XLSX.
    wb = _open_template(MBA_TEMPLATE_PATH)
    ws = wb['Flying Upload POD']
    # Clear any pre-existing data rows in the stub (row 2+).
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    rows: list[list[Any]] = []
    for ctx in contexts:
        row = _build_mba_row(
            safe_file_name=ctx['safe_file_name'],
            mba_listing=ctx['mba_listing'],
            global_listing=ctx['global_listing'],
            brand_name=ctx['brand_name'],
            category=ctx['category'],
            product_entry=ctx['product_entry'],
            workspace_id=ctx['workspace_id'],
            design_id=ctx['design_id'],
            summary=summary,
        )
        if row is None:
            continue
        rows.append(row)
    _write_mba_rows_to_sheet(ws, rows)
    summary.ready_rows = len(rows)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    zip_bytes = _pack_zip(xlsx_bytes, manifest)
    return zip_bytes, summary.as_dict()


def build_basic_bundle(
    workspace_id: str, design_ids: list[str],
) -> tuple[bytes, dict[str, Any]]:
    """Build Basic ZIP bundle (AC-95 + AC-96 + AC-97)."""
    ids = _validate_ids(design_ids)
    summary = _Summary(template='basic')
    contexts, manifest = _iterate_basic_context(workspace_id, ids, summary)

    wb = _open_template(BASIC_TEMPLATE_PATH)
    ws = wb['Flying Upload POD']
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    rows: list[list[Any]] = []
    for ctx in contexts:
        row = _build_basic_row(
            safe_file_name=ctx['safe_file_name'],
            global_listing=ctx['global_listing'],
        )
        rows.append(row)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    summary.ready_rows = len(rows)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    zip_bytes = _pack_zip(xlsx_bytes, manifest)
    return zip_bytes, summary.as_dict()


# ---------------------------------------------------------------------------
# CSV builders (AC-136)
# ---------------------------------------------------------------------------


def _csv_write_rows(
    header: list[str], rows: list[list[Any]],
) -> bytes:
    """UTF-8 with BOM, RFC 4180 quoted-CSV."""
    out = io.StringIO()
    writer = csv.writer(out, quoting=csv.QUOTE_ALL, lineterminator='\r\n')
    writer.writerow(header)
    for row in rows:
        writer.writerow(['' if v is None else v for v in row])
    return b'\xef\xbb\xbf' + out.getvalue().encode('utf-8')


_MBA_HEADER = [
    'Image Path', '',
    'Title DE', 'Description DE', 'Tags DE',
    'Title FR', 'Description FR', 'Tags FR',
    'Title IT', 'Description IT', 'Tags IT',
    'Title ES', 'Description ES', 'Tags ES',
    'Title JP', 'Description JP', 'Tags JP',
    'Title EN', 'Description EN', 'Tags EN',
    'Type', 'Color', '',
    'Brand DE', 'Brand FR', 'Brand IT', 'Brand ES', 'Brand JP', 'Brand EN',
    'Bullet 1 DE', 'Bullet 1 FR', 'Bullet 1 IT', 'Bullet 1 ES',
    'Bullet 1 JP', 'Bullet 1 EN',
    'Bullet 2 DE', 'Bullet 2 FR', 'Bullet 2 IT', 'Bullet 2 ES',
    'Bullet 2 JP', 'Bullet 2 EN',
    'Color1', 'Color2', 'Color3', 'Color4', 'Color5',
    'Color6', 'Color7', 'Color8', 'Color9', 'Color10',
    'Product', 'Marketplace',
    'Price US', 'Price UK', 'Price DE', 'Price FR',
    'Price IT', 'Price ES', 'Price JP',
    'Print', 'Draft', '',
    'Collection', 'Category', 'Background Color (Hex)',
]
assert len(_MBA_HEADER) == _MBA_HEADER_COUNT, (
    f'MBA header count drift: {len(_MBA_HEADER)} != {_MBA_HEADER_COUNT}'
)

_BASIC_HEADER = [
    'Image Path',
    'Title DE', 'Description DE', 'Tags DE',
    'Title EN', 'Description EN', 'Tags EN',
    'Type', 'Color',
]
assert len(_BASIC_HEADER) == _BASIC_HEADER_COUNT


def _bare_file_name(safe_name: str) -> str:
    """Strip ``designs/`` prefix for CSV Image Path (AC-136)."""
    return safe_name


def build_mba_csv(
    workspace_id: str, design_ids: list[str],
) -> tuple[bytes, dict[str, Any]]:
    """CSV variant of the MBA export (AC-136).

    Image Path column is the bare filename -- no ``designs/`` prefix.
    """
    ids = _validate_ids(design_ids)
    summary = _Summary(template='mba')
    contexts, _manifest = _iterate_mba_context(workspace_id, ids, summary)

    rows: list[list[Any]] = []
    for ctx in contexts:
        row = _build_mba_row(
            safe_file_name=ctx['safe_file_name'],
            mba_listing=ctx['mba_listing'],
            global_listing=ctx['global_listing'],
            brand_name=ctx['brand_name'],
            category=ctx['category'],
            product_entry=ctx['product_entry'],
            workspace_id=ctx['workspace_id'],
            design_id=ctx['design_id'],
            summary=summary,
        )
        if row is None:
            continue
        # AC-136: CSV Image Path = bare filename (no `designs/`).
        row[0] = _bare_file_name(ctx['safe_file_name'])
        rows.append(row)
    summary.ready_rows = len(rows)

    csv_bytes = _csv_write_rows(_MBA_HEADER, rows)
    return csv_bytes, summary.as_dict()


def build_basic_csv(
    workspace_id: str, design_ids: list[str],
) -> tuple[bytes, dict[str, Any]]:
    ids = _validate_ids(design_ids)
    summary = _Summary(template='basic')
    contexts, _manifest = _iterate_basic_context(workspace_id, ids, summary)

    rows: list[list[Any]] = []
    for ctx in contexts:
        row = _build_basic_row(
            safe_file_name=ctx['safe_file_name'],
            global_listing=ctx['global_listing'],
        )
        row[0] = _bare_file_name(ctx['safe_file_name'])
        rows.append(row)
    summary.ready_rows = len(rows)

    csv_bytes = _csv_write_rows(_BASIC_HEADER, rows)
    return csv_bytes, summary.as_dict()


# ---------------------------------------------------------------------------
# Preflight (AC-91)
# ---------------------------------------------------------------------------


def preflight(
    workspace_id: str,
    design_ids: list[str],
    template: str,
    fmt: str = 'xlsx',
) -> dict[str, Any]:
    """Run the same walking logic as ``build_*_bundle`` without writing bytes.

    Returns the summary dict. Raises ``ExportError`` on guardrail violations.
    """
    ids = _validate_ids(design_ids)
    summary = _Summary(template=template)

    if template == 'mba':
        contexts, _m = _iterate_mba_context(workspace_id, ids, summary)
        # Apply per-row validation via the row builder (detects unknown
        # products + "both" side + >10 colors warnings).
        ready = 0
        for ctx in contexts:
            row = _build_mba_row(
                safe_file_name=ctx['safe_file_name'],
                mba_listing=ctx['mba_listing'],
                global_listing=ctx['global_listing'],
                brand_name=ctx['brand_name'],
                category=ctx['category'],
                product_entry=ctx['product_entry'],
                workspace_id=ctx['workspace_id'],
                design_id=ctx['design_id'],
                summary=summary,
            )
            if row is not None:
                ready += 1
        summary.ready_rows = ready
    elif template == 'basic':
        contexts, _m = _iterate_basic_context(workspace_id, ids, summary)
        summary.ready_rows = len(contexts)
    else:
        raise ExportError(
            'invalid_template',
            f'Unknown template {template!r}.',
        )

    # format is accepted for parity with the public endpoint but does not
    # change the summary shape (AC-137).
    _ = fmt
    return summary.as_dict()
